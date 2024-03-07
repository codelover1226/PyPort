from flask import (
    Flask,
    render_template,
    url_for,
    request,
    redirect,
    session,
    jsonify,
    Response,
    json,
    send_file,
    send_from_directory,
    make_response,
    flash,
    abort,
    g
)
import re
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, desc, asc, and_, cast, Date
from flask_migrate import Migrate
from models import (
    db,
    User,
    Project,
    Vote,
    Comment,
    ProjectView,
    Bookmark,
    Report,
    WebsiteViews,
)
from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    login_required,
    current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
from authlib.integrations.flask_client import OAuth
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta, date
from werkzeug.middleware.proxy_fix import ProxyFix
from forms import RegistrationForm, LoginForm, CommentForm
from random import randint
from urllib.parse import quote, unquote
from markupsafe import Markup
from twilio.rest import Client
import matplotlib.pyplot as plt
from openpyxl import Workbook
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from collections import Counter
from bs4 import BeautifulSoup
import logging
import shutil
from PIL import Image
from pathlib import Path
from collections import Counter
import os
import pandas as pd
import random
import string
import json
import zipfile
import pytz
import io
import time
import threading
from io import StringIO
import random

from dotenv import load_dotenv

load_dotenv()

# Twilio credentials
account_sid = os.environ.get("TWILIO_ACCOUNT_SID")
auth_token = os.environ.get("TWILIO_AUTH_TOKEN")
twilio_number = os.environ.get("TWILIO_PHONE_NUMBER")


# Initialize the Flask app
app = Flask(__name__)
app.wsgi_app = ProxyFix(
    app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1
)
app.secret_key = "FREELANCER"
oauth = OAuth(app)

# Configure the database
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///project_voting.db"
# Define the UPLOAD_FOLDER
app.config[
    "UPLOAD_FOLDER"
] = "static/usersubmissions"  # Specify the folder where uploaded files will be saved


db.init_app(app)
migrate = Migrate(app, db)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# Configure logging
logging.basicConfig(level=logging.DEBUG)

ip_last_posted = {}
ip_last_submitted_project = {}
ip_last_added_marker = {}
ip_marker_additions = {}  # Initialize the dictionary to track marker additions
ip_project_submissions = {}


google = oauth.register(
    "google",
    client_id="FREELANCER",
    client_secret="FREELANCER",
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)


@app.before_request
def before_request():
    g.metaData = {
        'og_url': 'https://www.stimmungskompass.at/',
        'og_title': 'Stimmungskompass - Eine Plattform zur Bürgerbeteiligung',
        'og_description': 'Eine Plattform zur Bürgerbeteiligung. Engagiere dich für eine bessere Stadt!',
        'og_image': 'https://www.stimmungskompass.at/static/facebook_card.png'
    }

@app.route('/images/<path:filename>')
def serve_image(filename):
    image_dir = os.path.join(app.root_path, 'static')
    image_path = os.path.join(image_dir, filename)
    
    if 'image/webp' in request.headers.get('Accept', ''):
        webp_path = os.path.splitext(image_path)[0] + '.webp'
        
        if not os.path.exists(webp_path):
            os.makedirs(os.path.dirname(webp_path), exist_ok=True)
            try:
                with Image.open(image_path) as img:
                    # Adjust the quality parameter here (e.g., quality=75 for 75% quality)
                    img.save(webp_path, 'WEBP', quality=75)
            except IOError:
                abort(404)  # Image not found or unable to convert
        
        return send_from_directory(os.path.dirname(webp_path), os.path.basename(webp_path), mimetype='image/webp')
    else:
        return send_from_directory(os.path.dirname(image_path), os.path.basename(filename))


@app.route("/log_view")
def log_view():
    ip_address = request.remote_addr  # Get the IP address of the visitor

    # Check if this IP address has already been logged today
    existing_view = WebsiteViews.query.filter_by(
        view_date=datetime.utcnow().date(), ip_address=ip_address
    ).first()

    unique_views_count = WebsiteViews.query.filter_by(
        view_date=datetime.utcnow().date()
    ).count()

    if existing_view:
        # IP address has already been logged today
        return (
            jsonify(
                {
                    "message": f"IP {ip_address} not logged, it has already been logged today. There are {unique_views_count} unique viewers today."
                }
            ),
            200,
        )
    else:
        WebsiteViews.add_view(ip_address)
        print(f"New view logged for IP {ip_address}")  # Debugging
        return (
            jsonify(
                {
                    "message": f"IP {ip_address} logged. It is today the {unique_views_count + 1}th viewer."
                }
            ),
            200,
        )


@app.route("/get_unique_viewers_data")
def get_unique_viewers_data():
    try:
        start_date = request.args.get(
            "start_date", (datetime.utcnow() - timedelta(days=6)).strftime("%Y-%m-%d")
        )
        end_date = request.args.get("end_date", datetime.utcnow().strftime("%Y-%m-%d"))

        # Convert string dates to datetime.date objects
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()

        unique_viewers_data = (
            db.session.query(
                WebsiteViews.view_date,
                func.count(WebsiteViews.ip_address).label("viewer_count"),
            )
            .filter(WebsiteViews.view_date.between(start_date_obj, end_date_obj))
            .group_by(WebsiteViews.view_date)
            .all()
        )

        print("Queried Data:", unique_viewers_data)  # Debugging

        viewer_count_dict = {
            view_date.strftime("%Y-%m-%d"): viewer_count
            for view_date, viewer_count in unique_viewers_data
        }
        labels, values = zip(
            *[
                (
                    date.strftime("%Y-%m-%d"),
                    viewer_count_dict.get(date.strftime("%Y-%m-%d"), 0),
                )
                for date in (
                    start_date_obj + timedelta(n)
                    for n in range((end_date_obj - start_date_obj).days + 1)
                )
            ]
        )

        return jsonify({"labels": labels, "values": values})
    except Exception as e:
        print("Error:", str(e))  # More detailed error logging
        return jsonify({"error": str(e)}), 500


@app.route("/get_unique_viewers_data_range", methods=["POST"])
def get_unique_viewers_data_range():
    try:
        # Extract start and end dates from the request
        data = request.get_json()
        start_date = datetime.strptime(data["start_date"], "%Y-%m-%d").date()
        end_date = datetime.strptime(data["end_date"], "%Y-%m-%d").date()

        print("Startdatum:", start_date, "Enddatum:", end_date)  # Debugging

        # Query for viewer counts between the specified dates
        unique_viewers_data = (
            db.session.query(
                WebsiteViews.view_date,
                func.count(WebsiteViews.ip_address).label("viewer_count"),
            )
            .filter(WebsiteViews.view_date.between(start_date, end_date))
            .group_by(WebsiteViews.view_date)
            .all()
        )

        print("Queried Data:", unique_viewers_data)  # Debugging

        viewer_count_dict = {
            view_date.strftime("%Y-%m-%d"): viewer_count
            for view_date, viewer_count in unique_viewers_data
        }
        labels, values = zip(
            *[
                (
                    date.strftime("%Y-%m-%d"),
                    viewer_count_dict.get(date.strftime("%Y-%m-%d"), 0),
                )
                for date in (
                    start_date + timedelta(n)
                    for n in range((end_date - start_date).days + 1)
                )
            ]
        )

        return jsonify({"labels": labels, "values": values})
    except Exception as e:
        print("Error:", str(e))  # More detailed error logging
        return jsonify({"error": str(e)}), 500


@app.route("/project_submission_stats")
def project_submission_stats():
    try:
        start_date = request.args.get(
            "start", (date.today() - timedelta(days=7)).isoformat()
        )
        end_date = request.args.get("end", date.today().isoformat())
        include_map_objects = request.args.get("includeMapObjects") == "true"
        exclude_map_objects = request.args.get("excludeMapObjects") == "true"

        app.logger.debug(
            f"Fetching stats from {start_date} to {end_date}, include Notizens: {include_map_objects}, exclude Notizens: {exclude_map_objects}"
        )

        query = db.session.query(
            func.strftime("%Y-%m-%d", Project.date).label("submission_date"),
            Project.category,
            func.count(Project.id).label("project_count"),
        ).filter(Project.date.between(start_date, end_date))

        if include_map_objects:
            query = query.filter(Project.is_mapobject == True)
        elif exclude_map_objects:
            query = query.filter(Project.is_mapobject == False)

        submission_stats = query.group_by("submission_date", Project.category).all()

        app.logger.debug(f"Raw data: {submission_stats}")

        categories = set(category for _, category, _ in submission_stats)
        stats = {category: {} for category in categories}

        for submission_date, category, project_count in submission_stats:
            stats[category][submission_date] = project_count

        app.logger.debug(f"Processed stats: {stats}")
        return jsonify(stats)
    except Exception as e:
        app.logger.error(f"Error in project_submission_stats: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/get_activity_data")
def get_activity_data():
    start_date = request.args.get(
        "start", (date.today() - timedelta(days=7)).isoformat()
    )
    end_date = request.args.get("end", date.today().isoformat())

    # Fetching upvotes, downvotes, and comments
    upvotes = (
        db.session.query(
            func.date(Vote.timestamp).label("date"), func.count().label("count")
        )
        .filter(Vote.timestamp.between(start_date, end_date), Vote.upvote == True)
        .group_by(func.date(Vote.timestamp))
        .all()
    )

    downvotes = (
        db.session.query(
            func.date(Vote.timestamp).label("date"), func.count().label("count")
        )
        .filter(Vote.timestamp.between(start_date, end_date), Vote.downvote == True)
        .group_by(func.date(Vote.timestamp))
        .all()
    )

    comments = (
        db.session.query(
            func.date(Comment.timestamp).label("date"), func.count().label("count")
        )
        .filter(Comment.timestamp.between(start_date, end_date))
        .group_by(func.date(Comment.timestamp))
        .all()
    )

    # Fetching reports
    reports = (
        db.session.query(
            func.date(Report.timestamp).label("date"), func.count().label("count")
        )
        .filter(Report.timestamp.between(start_date, end_date))
        .group_by(func.date(Report.timestamp))
        .all()
    )

    # Fetching bookmarks
    bookmarks = (
        db.session.query(
            func.date(Bookmark.timestamp).label("date"), func.count().label("count")
        )
        .filter(Bookmark.timestamp.between(start_date, end_date))
        .group_by(func.date(Bookmark.timestamp))
        .all()
    )

    # Convert query results to dictionaries
    upvotes_dict = {str(day.date): day.count for day in upvotes}
    downvotes_dict = {str(day.date): day.count for day in downvotes}
    comments_dict = {str(day.date): day.count for day in comments}
    reports_dict = {str(day.date): day.count for day in reports}
    bookmarks_dict = {str(day.date): day.count for day in bookmarks}

    activity_data = {
        "upvotes": upvotes_dict,
        "downvotes": downvotes_dict,
        "comments": comments_dict,
        "reports": reports_dict,
        "bookmarks": bookmarks_dict,
    }

    return jsonify(activity_data)


@app.route("/get_view_data")
def get_view_data():
    start_date = request.args.get(
        "start", (datetime.today() - timedelta(days=7)).strftime("%Y-%m-%d")
    )
    end_date = request.args.get("end", datetime.today().strftime("%Y-%m-%d"))

    daily_views = (
        db.session.query(
            func.date(Project.date).label("date"),
            func.sum(Project.view_count).label("daily_view_count"),
        )
        .filter(Project.date.between(start_date, end_date))
        .group_by(func.date(Project.date))
        .all()
    )

    total_views = (
        db.session.query(func.sum(Project.view_count))
        .filter(Project.date.between(start_date, end_date))
        .scalar()
    )

    daily_views_dict = {str(day.date): day.daily_view_count for day in daily_views}

    view_data = {"daily_views": daily_views_dict, "total_views": total_views}

    return jsonify(view_data)


@app.route("/get_engaged_projects")
@login_required
def get_engaged_projects():
    threshold = request.args.get("threshold", default=1, type=int)

    # Query projects with view_count greater than or equal to threshold
    projects = Project.query.filter(Project.view_count >= threshold).all()

    project_list = []
    for project in projects:
        project_details = {
            "id": project.id,
            "name": project.name,
            "category": project.category,  # Include the category here
            "views": project.view_count,
            "upvotes": len([vote for vote in project.votes if vote.upvote]),
            "downvotes": len([vote for vote in project.votes if vote.downvote]),
            "comments": len(project.comments),
            "bookmarks": len(Bookmark.query.filter_by(project_id=project.id).all()),
            "reports": project.p_reports,
        }
        project_list.append(project_details)

    return jsonify(project_list)


@app.route("/get_chart_data", methods=["GET"])
@login_required
def get_chart_data():
    filter_param = request.args.get("filter", "all")
    description_length_filter = request.args.get("description_length_filter", "all")

    projects_query = Project.query
    description_projects_query = Project.query

    if filter_param == "true":
        projects_query = projects_query.filter_by(is_mapobject=True)
    elif filter_param == "false":
        projects_query = projects_query.filter_by(is_mapobject=False)

    if description_length_filter == "true":
        description_projects_query = description_projects_query.filter_by(
            is_mapobject=True
        )
    elif description_length_filter == "false":
        description_projects_query = description_projects_query.filter_by(
            is_mapobject=False
        )

    # Count projects per category with the filtered query
    category_counts = (
        projects_query.with_entities(Project.category, func.count(Project.id))
        .group_by(Project.category)
        .all()
    )
    category_counts = {category: count for category, count in category_counts}

    # Count Notizens vs Projektvorschläge
    mapobject_counts = {
        "Notizen": Project.query.filter_by(is_mapobject=True).count(),
        "Projektvorschläge": Project.query.filter_by(is_mapobject=False).count(),
    }

    description_length_ranges = ["0-50", "51-100", "101-150", "151-200", ">200"]
    description_length_counts = {range: 0 for range in description_length_ranges}
    total_description_length = 0
    description_projects_count = description_projects_query.count()

    for project in description_projects_query:
        length = len(project.descriptionwhy)
        total_description_length += length
        if length <= 50:
            description_length_counts["0-50"] += 1
        elif 51 <= length <= 100:
            description_length_counts["51-100"] += 1
        elif 101 <= length <= 150:
            description_length_counts["101-150"] += 1
        elif 151 <= length <= 200:
            description_length_counts["151-200"] += 1
        else:
            description_length_counts[">200"] += 1

    average_description_length = (
        (total_description_length / description_projects_count)
        if description_projects_count > 0
        else 0
    )

    return jsonify(
        {
            "categoryCounts": category_counts,
            "mapobjectCounts": mapobject_counts,
            "descriptionLengthCounts": description_length_counts,
            "averageDescriptionLength": average_description_length,
        }
    )


def get_project_category_chart_data():
    # Count projects per category
    projects = Project.query.all()
    category_counts = Counter([project.category for project in projects])

    # Prepare data for the pie chart
    pie_chart_data = {
        "labels": list(category_counts.keys()),
        "datasets": [
            {
                "label": "Project Categories",
                "data": list(category_counts.values()),
                "backgroundColor": [
                    "#ff6384",
                    "#36a2eb",
                    "#cc65fe",
                    "#ffce56",
                    "#4bc0c0",
                    "#f77825",
                ],  # Customize colors
            }
        ],
    }
    return pie_chart_data


@app.route("/login/google")
def google_login():
    # Generate a nonce and store it in the session
    nonce = "".join(random.choices(string.ascii_uppercase + string.digits, k=16))
    session["google_auth_nonce"] = nonce

    redirect_uri = url_for("authorized", _external=True)
    return oauth.google.authorize_redirect(
        redirect_uri, nonce=nonce, prompt="select_account"
    )


def generate_pie_chart_categories(workbook, categories):
    logging.debug("Generating pie chart for categories")
    pie_sheet = workbook.create_sheet(title="Categories Pie Chart")
    for row in categories.items():
        pie_sheet.append(row)
    chart = PieChart()
    labels = Reference(pie_sheet, min_col=1, min_row=2, max_row=len(categories) + 1)
    data = Reference(pie_sheet, min_col=2, min_row=1, max_row=len(categories) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    pie_sheet.add_chart(chart, "E2")


# Function to generate pie chart for description lengths
def generate_pie_chart_description_length(workbook, length_ranges):
    logging.debug("Generating pie chart for description lengths")
    pie_sheet = workbook.create_sheet(title="Description Lengths Pie Chart")
    for row in length_ranges.items():
        pie_sheet.append(row)
    chart = PieChart()
    labels = Reference(pie_sheet, min_col=1, min_row=2, max_row=len(length_ranges) + 1)
    data = Reference(pie_sheet, min_col=2, min_row=1, max_row=len(length_ranges) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    pie_sheet.add_chart(chart, "E2")


def calculate_votes(project_id):
    project = Project.query.get(project_id)
    if project:
        upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
        downvotes = Vote.query.filter_by(project_id=project.id, downvote=True).count()
        # upvotes = sum(1 for vote in project.votes if vote.upvote)
        # downvotes = sum(1 for vote in project.votes if vote.downvote)
        return upvotes, downvotes
    else:
        return 0, 0


@app.route("/export_gis", methods=["GET"])
@login_required
def export_gis():
    category = request.args.get("category", "")

    if category and category != "Alle Kategorien":
        projects = Project.query.filter_by(category=category).all()
    else:
        projects = Project.query.all()

    features = []
    for project in projects:
        if project.geoloc and "," in project.geoloc:
            lat, lon = project.geoloc.split(",")
            try:
                lat, lon = float(lat.strip()), float(lon.strip())
                features.append(
                    {
                        "type": "Feature",
                        "geometry": {"type": "Point", "coordinates": [lon, lat]},
                        "properties": project.to_dict(),
                    }
                )
            except ValueError:
                pass  # Handle invalid geolocation data

    geojson_data = {"type": "FeatureCollection", "features": features}

    formatted_geojson = json.dumps(geojson_data, indent=4)
    response = Response(
        formatted_geojson,
        mimetype="application/json",
        headers={"Content-Disposition": "attachment;filename=exported_data.geojson"},
    )
    return response


@app.route("/export_csv", methods=["GET", "POST"])
@login_required
def export_csv():
    category = request.args.get("category", "")

    # Querying the database based on the category
    if category and category != "Alle Kategorien":
        projects_data = Project.query.filter_by(category=category).all()
    else:
        projects_data = Project.query.all()

    projects_list = [project.to_dict() for project in projects_data]

    # Convert to DataFrame
    df = pd.DataFrame(projects_list)

    # Converting DataFrame to CSV
    csv_data = StringIO()
    df.to_csv(csv_data, index=False)

    # Flask response
    response = make_response(csv_data.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=projects_data.csv"
    response.headers["Content-type"] = "text/csv"


    return response

@app.route("/export_projects", methods=["GET", "POST"])
@login_required
def export_projects():
    try:
        category = request.values.get("category")       
        include_comments = "include_comments" in request.values
        include_votes = "include_votes" in request.values
        query = Project.query
        if category:
            query = query.filter_by(category=category)
        projects = query.all()
        # Convert to DataFrame and strip HTML tags from specific fields
        def strip_html(content):
            if content:
                return BeautifulSoup(content, "html.parser").get_text()
            return content
        projects_data = []
        for project in projects:
            app.logger.info(project)
            project_dict = project.to_dict()
            # Update fields with HTML content
            project_dict["descriptionwhy"] = strip_html(project_dict["descriptionwhy"])
            project_dict["public_benefit"] = strip_html(
                project_dict["public_benefit"]
            )  # Apply to 'public_benefit'

            # Calculate upvotes and downvotes
            upvotes, downvotes = calculate_votes(project.id)
            print(
                f"Project ID {project.id} exported with {upvotes} upvotes and {downvotes} downvotes into the Excel table"
            )
            project_dict["upvotes"] = upvotes
            project_dict["downvotes"] = downvotes

            projects_data.append(project_dict)

        df = pd.DataFrame(projects_data)
        rename_columns = {
            "id": "ID",
            "category": "Kategorie",
            "name": "Titel",
            "descriptionwhy": "Beschreibung",
            "public_benefit": "Vorteile",
            "date": "Datum",
            "geoloc": "Markierter Standort",
            "author": "Author",
            "image_file": "Bild",
            "is_important": "Privat markiert",
            "is_featured": "Ausgewählt",
            "view_count" : "View Count",
            "p_reports" : "P Reports",
            "upvotes": "Upvotes",  # New columns for upvotes and downvotes
            "downvotes": "Downvotes",
        }
        df = df.rename(columns=rename_columns)[
            list(rename_columns.values())
        ]  # Reorder columns
        app.logger.info(df.to_json)
        app.logger.info(projects)

        # Remove the 'p_reports' column
        # df.drop(columns=["p_reports"], inplace=True, errors="ignore")

        def format_geoloc(geoloc):
            try:
                if geoloc and "," in geoloc:
                    lat, lon = geoloc.split(",")
                    return f"https://www.google.com/maps/search/?api=1&query={lat.strip()},{lon.strip()}"
                else:
                    return ""
            except Exception as e:
                app.logger.error(f"Error in format_geoloc: {e}")
                return ""

        # Check if 'geoloc' column exists in DataFrame
        if "geoloc" in df.columns:
            df["geoloc"] = df["geoloc"].apply(format_geoloc)
        else:
            app.logger.warning("'geoloc' column not found in DataFrame")

        if include_comments or include_votes:
            for project in projects:
                # Include comments and votes if requested
                if include_comments:
                    # Add comment data
                    project_comments = Comment.query.filter_by(
                        project_id=project.id
                    ).all()
                    comments_data = [
                        {
                            "project_id": project.id,
                            "comment_id": comment.id,
                            "comment_text": comment.text,
                            "comment_author_id": comment.user_id,
                            "comment_timestamp": comment.timestamp.strftime(
                                "%Y-%m-%d %H:%M:%S"
                            ),  # Format timestamp
                        }
                        for comment in project_comments
                    ]
                    comments_df = pd.DataFrame(comments_data)
                    df = pd.merge(
                        df, comments_df, how="left", left_on="id", right_on="project_id"
                    )

                    

        filename = "exported_projects.xlsx"
        filepath = os.path.join("static/excel", filename)
        writer = pd.ExcelWriter(filepath, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Exported Projects")
        workbook = writer.book
        worksheet = writer.sheets["Exported Projects"]

                    
        # Medium border style
        medium_border_side = Side(border_style="medium", color="000000")
        medium_border = Border(top=medium_border_side, bottom=medium_border_side)

        # Font for all cells
        font_size = 12  # Set your desired font size
        bahnschrift_font = Font(name="Bahnschrift", size=font_size)
        # Formatting for the first row
        header_font = Font(
            name="Bahnschrift", bold=True, color="F5F1E4", size=font_size + 2
        )  # Slightly larger for headers
        header_fill = PatternFill(
            start_color="003056", end_color="003056", fill_type="solid"
        )
        kategorie_font = Font(name="Bahnschrift", size=12, bold=True)

        # Apply font settings to Kategorie column
        kategorie_col_idx = df.columns.get_loc("Kategorie") + 1  # 1-based indexing
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=kategorie_col_idx,
            max_col=kategorie_col_idx,
        ):
            for cell in row:
                cell.font = kategorie_font

        for row in worksheet.iter_rows(
            min_row=2, min_col=2, max_col=2
        ):  # Kategorie is the 2nd column
            for cell in row:
                cell.font = Font(name="Bahnschrift", size=12, bold=True)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = bahnschrift_font  # Set font for all cells
                cell.border = medium_border  # Set medium border for all cells
                if cell.row == 1:
                    cell.font = header_font  # Override font for the first row
                    cell.fill = header_fill  # Apply fill for the first row
                    cell.alignment = Alignment(horizontal="center", vertical="center")


        colors = ["F5F1E4", "D9D4C7"]

        for i, column_cells in enumerate(worksheet.columns):
            color_index = i % len(colors)  # Alternate between 0 and 1
            fill = PatternFill(
                start_color=colors[color_index],
                end_color=colors[color_index],
                fill_type="solid",
            )
            for cell in column_cells[1:]:  # Skip the first row
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths based on the longest content
        max_char_length = 100
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            length = min(length, max_char_length)  # Limit to max_char_length characters
            col_width = length * 1.3  # Approximate column width
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = col_width

            # Pie Chart for Project Categories
        start_row_for_charts = df.shape[0] + 3  # 2 rows gap after the DataFrame

        # Pie Chart for Project Categories
        # app.logger.debug("Generating pie chart for project categories")
        category_counts = Counter(df["Kategorie"])
        fig, ax = plt.subplots()
        ax.pie(
            category_counts.values(), labels=category_counts.keys(), autopct="%1.1f%%"
        )
        plt.title("Project Categories Distribution")
        img_data = io.BytesIO()
        plt.savefig(img_data, format="png")
        img_data.seek(0)
        app.logger.info("here")
        img = ExcelImage(img_data)
        chart1_cell = f"D{start_row_for_charts}"  # Adjust as needed
        worksheet.add_image(img, chart1_cell)
        app.logger.debug("Pie chart for project categories created")

        # Pie Chart for Average Description Length
        app.logger.debug("Generating pie chart for average description length")

        def get_length_category(row):
            total_length = len(row["Beschreibung"]) + len(row["Vorteile"])
            if 50 <= total_length < 100:
                return "50-100"
            elif 100 <= total_length < 200:
                return "100-200"
            elif 200 <= total_length < 400:
                return "200-400"
            elif 400 <= total_length < 800:
                return "400-800"
            else:
                return "800-5000"
        df["Length Category"] = df.apply(get_length_category, axis=1)
        length_category_counts = Counter(df["Length Category"])
        fig, ax = plt.subplots()
        ax.pie(
            length_category_counts.values(),
            labels=length_category_counts.keys(),
            autopct="%1.1f%%",
        )
        plt.title("Description Length Distribution")
        img_data = io.BytesIO()
        plt.savefig(img_data, format="png")
        img_data.seek(0)
        img = ExcelImage(img_data)
        chart2_cell = f"f{start_row_for_charts}"  # Adjust as needed
        worksheet.add_image(img, chart2_cell)
        app.logger.debug("Pie chart for average description length created")

        mapobject_counts = {
            "Notizen": Project.query.filter_by(is_mapobject=True).count(),
            "Projektvorschläge": Project.query.filter_by(is_mapobject=False).count(),
        }

        # Generate the pie chart for Notizen vs Projektvorschläge
        app.logger.debug("Generating pie chart for Notizen vs Projektvorschläge")
        fig, ax = plt.subplots()
        ax.pie(
            mapobject_counts.values(), labels=mapobject_counts.keys(), autopct="%1.1f%%"
        )
        plt.title("Ratio Projekt vs Map_object")
        img_data = io.BytesIO()
        plt.savefig(img_data, format="png")
        img_data.seek(0)
        img = ExcelImage(img_data)
        chart3_cell = f"E{ start_row_for_charts}"  # Adjust the cell location as needed
        worksheet.add_image(img, chart3_cell)
        app.logger.debug("Pie chart for Notizen vs Projektvorschläge created")

        writer._save()
        app.logger.debug(f"Excel file with pie charts saved at {filepath}")

        return jsonify({"filepath": filepath})
    except Exception as e:
        print("Error:", str(e))  # More detailed error logging
        return jsonify({"error": str(e)}), 500

@app.route("/get_categories")
def get_categories():
    try:
        categories = [
            project.category
            for project in Project.query.distinct(Project.category).all()
        ]
        app.logger.debug("Fetched categories for dropdown")
        return jsonify(success=True, categories=categories)
    except Exception as e:
        app.logger.error(f"Error in get_categories: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/get_unique_categories")
def get_unique_categories():
    try:
        categories = [
            category[0]
            for category in db.session.query(Project.category.distinct()).all()
        ]
        app.logger.debug("Fetched unique categories for dropdown")
        return jsonify(success=True, categories=categories)
    except Exception as e:
        app.logger.error(f"Error in get_unique_categories: {e}")
        return jsonify(success=False, error=str(e)), 500


@app.route("/login/google/authorized")
def authorized():
    token = oauth.google.authorize_access_token()

    nonce = session.pop("google_auth_nonce", None)
    user_info = oauth.google.parse_id_token(token, nonce=nonce)

    existing_user = User.query.filter_by(phone_number=user_info.get("email")).first()

    if not existing_user:
        # Convert UTC time to desired timezone
        tz = pytz.timezone(
            "Europe/Berlin"
        )  # Replace 'Europe/Berlin' with your desired timezone
        account_creation_time = (
            datetime.utcnow().replace(tzinfo=pytz.utc).astimezone(tz)
        )

        new_user = User(
            phone_number=user_info.get("email"),
            name=user_info.get("name", "Unknown"),
            account_creation=account_creation_time,
            is_googleaccount=True,
            is_admin=False,
            password_hash="google_oauth",  # Set a default value
        )
        db.session.add(new_user)
        db.session.commit()
        login_user(new_user)
    else:
        login_user(existing_user)

    return redirect(url_for("index"))


# Create the database tables before the first request
def create_tables():
    db.create_all()


def can_user_post_comment(user_id):
    time_limit = datetime.now() - timedelta(minutes=15)
    max_comments = 5
    recent_comments = (
        Comment.query.filter(Comment.user_id == user_id, Comment.timestamp > time_limit)
        .order_by(Comment.timestamp.asc())
        .all()
    )

    if len(recent_comments) >= max_comments:
        # Calculate the reset time as 15 minutes from the oldest comment in the last 15 minutes
        oldest_comment_time = recent_comments[0].timestamp
        reset_time = oldest_comment_time + timedelta(minutes=15)
        return False, reset_time
    else:
        return True, None


def get_reset_time(user_id):
    earliest_comment = (
        Comment.query.filter(Comment.user_id == user_id)
        .order_by(Comment.timestamp.asc())
        .first()
    )

    if earliest_comment:
        return earliest_comment.timestamp + timedelta(minutes=15)
    else:
        return datetime.now()


def zip_user_submissions():
    try:
        # Path for the directory to be zipped
        directory_to_zip = Path(app.root_path, "static", "usersubmissions")

        # Path for the zip file
        zip_path = Path(app.root_path, "static", "usersubmissions.zip")

        # Check if directory exists and has files
        if directory_to_zip.is_dir() and any(directory_to_zip.iterdir()):
            # Create a zip file
            shutil.make_archive(
                zip_path.stem, "zip", directory_to_zip.parent, directory_to_zip.name
            )
            logging.debug("Zip file created at: %s", zip_path)
            return zip_path
        else:
            logging.debug("No directory or files to zip")
            return None
    except Exception as e:
        logging.error("Error creating zip file: %s", e)
        return None


@app.route("/delete_all_projects", methods=["POST"])
@login_required
def delete_all_projects():
    # Check if the user is the admin
    if current_user.id != 1:
        return redirect(url_for("index"))

    try:
        # Fetch all projects
        projects = Project.query.all()

        # Löschen each project to trigger cascade deletion for comments
        for project in projects:
            db.session.delete(project)

        # Commit the changes to the database
        db.session.commit()

    except Exception as e:
        db.session.rollback()  # Rollback in case of error
      
    return redirect(url_for("admintools"))


@app.route("/download_images")
def download_images():
    try:
        zip_path = zip_user_submissions()
        if zip_path and zip_path.is_file():
            # Provide the full path to the file for direct download
            full_path = zip_path.resolve()

            # Create a JavaScript snippet to initiate the download
            download_script = f'<script>window.location.href = "{url_for("static", filename="usersubmissions.zip")}";</script>'

            # Return the script as an HTML response
            return Response(download_script, mimetype="text/html")
        else:
            flash('No images available to download.', 'info')
            return redirect(url_for("profil"))
    except Exception as e:
        logging.error("Error in downloading images: %s", e)
        flash('Error in downloading images.', 'danger')
        return redirect(url_for("profil"))


@app.route("/")
def index():
    projects = Project.query.all()
    featured_projects = Project.query.filter_by(is_featured=True).all()
    ip_address = request.remote_addr
    WebsiteViews.add_view(ip_address)

    # Calculate upvotes and downvotes for each project
    for project in projects + featured_projects:
        upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
        downvotes = Vote.query.filter_by(project_id=project.id, downvote=True).count()

        project.upvotes = upvotes
        project.downvotes = downvotes
        project.upvote_percentage = (
            (upvotes / (upvotes + downvotes) * 100) if (upvotes + downvotes) > 0 else 0
        )
        project.downvote_percentage = (
            (downvotes / (upvotes + downvotes) * 100)
            if (upvotes + downvotes) > 0
            else 0
        )

    # Count projects where is_mapobject is false
    project_count_non_map = Project.query.filter_by(is_mapobject=False).count()

    # Count projects where is_mapobject is true
    mapobject_count = Project.query.filter_by(is_mapobject=True).count()

    metaData = g.metaData


    return render_template(
        "index.html",
        projects=projects,
        project_count=project_count_non_map,
        mapobject_count=mapobject_count,
        featured_projects=featured_projects,
        metaData=metaData,
    )


@app.route("/logout")
@login_required
def logout():
    logout_user()
    next_page = request.args.get("next")
    return redirect(
        next_page or url_for("index")
    )  # 'index' is the function name of your home route


@app.route("/init_projects")
def init_projects():
    projects = [
        {
            "name": "Road improvements",
            "description": "Project includes: xyz",
            "image_file": "road.png",
        },
        {
            "name": "Park improvements",
            "description": "Project includes: xyz",
            "image_file": "park.png",
        },
        {
            "name": "New train station",
            "description": "Project includes: xyz",
            "image_file": "train station.png",
        },
        {
            "name": "New kindergarten",
            "description": "Project includes: xyz",
            "image_file": "kindergarten.png",
        },
    ]

    for proj in projects:
        if not Project.query.filter_by(name=proj["name"]).first():
            new_project = Project(
                name=proj["name"],
                description=proj["description"],
                image_file=proj["image_file"],
            )
            db.session.add(new_project)
            logging.debug("Adding project: %s", proj["name"])

    db.session.commit()
    return "Projects initialized successfully"


@app.route("/download/<filename>")
def download_file(filename):
    decoded_filename = unquote(filename)
    # Further processing with decoded_filename
    return send_from_directory(directory, decoded_filename)


def calculate_age(dob):
    today = datetime.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        logging.debug(
            "Registration attempt: %s", request.form
        )  # Debug: print form data
        ip_address = request.remote_addr
        WebsiteViews.add_view(ip_address)
        phone_number = request.form.get("phone_number")
        password = request.form.get("password")
        name = request.form.get("name")
        ip_address = request.remote_addr  # Capture IP address from request

        # Check for existing user with the same Ihre Handynummer
        existing_user = User.query.filter_by(phone_number=phone_number).first()
        if existing_user:
            flash('An account with this Ihre Handynummer already exists.', 'danger')
            logging.debug(
                "Account registration failed: Ihre Handynummer already exists"
            )
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Diese Handynummer ist bereits registriert.",
                    }
                ),
                400,
            )

        # Generate OTP and handle verification
        otp = randint(100000, 999999)
        client = Client(account_sid, auth_token)
        message = client.messages.create(
            body=f"Stimmungskompass: Ihr OTP ist: {otp}. Fügen Sie hinzu, um die Registrierung abzuschließen!",
            from_=twilio_number,
            to=phone_number,
        )
        logging.debug(f"Twilio message S{message.sid}")

        # Save the user data in session temporarily
        session["user_data"] = {
            "phone_number": phone_number,
            "password": password,
            "name": name,
            "ip_address": ip_address,
            "otp": otp,
        }

        logging.debug("OTP sent for verification to Ihre Handynummer: %s", phone_number)
        return jsonify({"success": True, "message": "OTP sent successfully"})
    metaData=g.metaData
    return render_template("register/index.html", metaData=metaData)


# Function to clean up old IP addresses
def cleanup_ip_addresses():
    while True:
        time.sleep(60)  # Check every minute
        current_time = datetime.now()

        # Cleanup for ip_last_submitted_project
        for ip in list(ip_last_submitted_project.keys()):
            if (
                current_time - ip_last_submitted_project[ip]
            ).seconds > 86400:  # 24 hours
                del ip_last_submitted_project[ip]
                print(f"Removed IP {ip} from submitted project tracking")

        # Cleanup for ip_last_added_marker
        for ip in list(ip_last_added_marker.keys()):
            if (current_time - ip_last_added_marker[ip]).seconds > 86400:  # 24 hours
                del ip_last_added_marker[ip]
                print(f"Removed IP {ip} from added marker tracking")


# Start the background thread for cleanup
cleanup_thread = threading.Thread(target=cleanup_ip_addresses, daemon=True)
cleanup_thread.start()


@app.route("/check_marker_limit", methods=["GET"])
def check_marker_limit():
    ip_address = request.remote_addr
    now = datetime.now()
    additions = ip_marker_additions.get(ip_address, [])
    additions = [time for time in additions if now - time < timedelta(days=1)]
    current_count = len(additions)
    max_limit = 10  # Set your max limit here
    limit_reached = current_count >= max_limit

    if limit_reached and additions:
        reset_time = max(additions) + timedelta(days=1)
    else:
        reset_time = None

    return jsonify(
        {
            "ip_address": ip_address,
            "current_count": current_count,
            "max_limit": max_limit,
            "limit_reached": limit_reached,
            "reset_time": reset_time.isoformat() if reset_time else None,
        }
    )


@app.route("/add_marker", methods=["POST"])
def add_marker():
    data = request.json
    ip_address = request.remote_addr

    # Check and update the rate limit for marker additions
    now = datetime.now()
    additions = ip_marker_additions.get(ip_address, [])
    # Filter additions within the last 24 hours
    additions = [time for time in additions if now - time < timedelta(days=1)]

    if len(additions) >= 300:  # Assuming a limit of 2 markers per day
        app.logger.warning(
            f"IP {ip_address} blocked from adding new markers due to rate limit"
        )
        return (
            jsonify(
                {
                    "error": "Rate limit exceeded. You can only add 2 markers every 24 hours"
                }
            ),
            429,
        )

    try:
        author_id = current_user.id if current_user.is_authenticated else 0
        public_benefit = data.get("public_benefit", "-")
        image_file = data.get("image_file", "keinbild.jpg")

        new_project = Project(
            name="Notiz",
            category=data["category"],
            descriptionwhy=data["description"],
            public_benefit=public_benefit,
            image_file=image_file,
            geoloc=f"{data['lat']}, {data['lng']}",
            author=author_id,
            is_mapobject=True,
        )

        db.session.add(new_project)
        db.session.commit()

        # Update IP tracking dictionary
        additions.append(now)
        ip_marker_additions[ip_address] = additions

        app.logger.info(f"IP {ip_address} recorded for posting a marker")
        return (
            jsonify({"message": "Marker added successfully", "id": new_project.id}),
            200,
        )

    except Exception as e:
        app.logger.error(f"Error saving marker: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/get_projects")
def get_projects():
    try:
        projects = Project.query.all()
        projects_data = []
        for project in projects:
            project_data = project.to_dict()
            project_data[
                "is_mapobject"
            ] = project.is_mapobject  # Include the is_mapobject attribute
            upvotes = sum(1 for vote in project.votes if vote.upvote)
            downvotes = sum(1 for vote in project.votes if vote.downvote)
            upvote_percentage = (
                (upvotes / (upvotes + downvotes) * 100)
                if (upvotes + downvotes) > 0
                else 0
            )
            downvote_percentage = (
                (downvotes / (upvotes + downvotes) * 100)
                if (upvotes + downvotes) > 0
                else 0
            )

            project_data["upvotes"] = upvotes
            project_data["downvotes"] = downvotes
            project_data["upvote_percentage"] = upvote_percentage
            project_data["downvote_percentage"] = downvote_percentage
            projects_data.append(project_data)
        return jsonify(projects_data)
    except Exception as e:
        logging.error(f"Error in get_projects: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/check_limit", methods=["GET"])
def check_limit():
    ip_address = request.remote_addr
    additions = ip_marker_additions.get(ip_address, [])

    # Filter additions within the last 24 hours
    additions = [
        time for time in additions if datetime.now() - time < timedelta(days=1)
    ]
    ip_marker_additions[ip_address] = additions  # Update the dictionary

    # Debugging log
    app.logger.debug(f"IP {ip_address} - Marker Additions: {additions}")

    limit_reached = len(additions) >= 10  # Assuming a limit of 10
    reset_time = max(additions) + timedelta(days=1) if limit_reached else None

    app.logger.debug(
        f"IP {ip_address} - Limit Reached: {limit_reached}, Reset Time: {reset_time}"
    )

    return jsonify(
        {
            "limit_reached": limit_reached,
            "reset_time": reset_time.isoformat() if reset_time else None,
        }
    )


@app.route("/check_project_limit")
def check_project_limit():
    ip_address = request.remote_addr
    submissions = ip_project_submissions.get(ip_address, [])

    # Filtern submissions within the last 24 hours
    submissions = [
        time for time in submissions if datetime.now() - time < timedelta(days=1)
    ]

    limit_reached = len(submissions) >= 5  # Assuming a limit of 5 submissions per day
    reset_time = (
        (submissions[0] + timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
        if limit_reached
        else None
    )

    # Log for debugging
    app.logger.debug(f"IP Address: {ip_address}")
    app.logger.debug(f"Submissions: {submissions}")
    app.logger.debug(f"Limit Reached: {limit_reached}")
    app.logger.debug(f"Reset Time: {reset_time}")

    return jsonify({"limit_reached": limit_reached, "reset_time": reset_time})


def generate_otp_and_send(phone_number):
    otp = randint(100000, 999999)
    client = Client(account_sid, auth_token)
    try:
        message = client.messages.create(
            body=f"Stimmungskompass: Ihr OTP ist: {otp}",
            from_=twilio_number,
            to=phone_number,
        )
        logging.debug(f"OTP sent: {message.sid}")
    except Exception as e:
        logging.error(f"Error sending OTP: {e}")
        return None
    return otp


@app.route("/favicon.ico")
def favicon():
    app.logger.debug("Favicon loaded successfully")  # Add this debug message
    return url_for("static", filename="favicon.ico")


@app.route("/karte")
def karte():
    # Fetch your projects data from the database or any source
    ip_address = request.remote_addr
    WebsiteViews.add_view(ip_address)
    projects = (
        get_projects()
    )  # This is a placeholder for your actual function to fetch projects
    metaData = g.metaData
    return render_template("karte/index.html", projects=projects, metaData=metaData)


@app.route("/request_otp", methods=["POST"])
def request_otp():
    phone_number = request.form.get("phone_number")
    user = User.query.filter_by(phone_number=phone_number).first()
    if user:
        otp = generate_otp_and_send(phone_number)
        if otp:
            session["reset_otp"] = otp
            session["phone_number"] = phone_number
            logging.debug(
                f"OTP generated and session updated for Ihre Handynummer {phone_number}"
            )
            return jsonify(success=True, message="OTP sent to your phone.")
        else:
            return jsonify(success=False, message="Failed to send OTP.")
    return jsonify(success=False, message="Ihre Handynummer not found.")


@app.route("/reset_password", methods=["GET", "POST"])
def reset_password():
    
    metaData=g.metaData
    if request.method == "POST":
        otp_entered = request.form.get("otp")
        new_password = request.form.get("new_password")
        phone_number = session.get("phone_number")
        if "reset_otp" in session and session["reset_otp"] == int(otp_entered):
            user = User.query.filter_by(phone_number=phone_number).first()
            if user:
                user.set_password(new_password)
                db.session.commit()
                logging.debug(
                    f"Password reset for user with Ihre Handynummer {phone_number}"
                )
                flash('Your password has been reset successfully.', 'success')
                return redirect(url_for("login"))
            # else:
            flash('Invalid Ihre Handynummer.', 'danger')
        # else:
        flash('Invalid OTP. Please try again.', 'danger')
    return render_template("reset_password.html", metaData=metaData)


@app.route("/neuerbeitrag")
def neuerbeitrag():
    metaData=g.metaData
    return render_template("neuerbeitrag/index.html", metaData=metaData)


@app.route("/submit_project", methods=["GET", "POST"])
@login_required
def submit_project():
    if request.method == "POST":
        ip_address = request.remote_addr
        ip_address = request.remote_addr
        WebsiteViews.add_view(ip_address)
        # Check and update the rate limit for project submissions
        submissions = ip_project_submissions.get(ip_address, [])
        # Remove timestamps older than 24 hours
        submissions = [
            time for time in submissions if datetime.now() - time < timedelta(days=1)
        ]
        if len(submissions) >= 5:
            app.logger.warning(
                f"IP {ip_address} blocked from submitting new projects due to rate limit"
            )
            return (
                jsonify(
                    {
                        "error": "Rate limit exceeded. You can only submit 5 projects every 24 hours"
                    }
                ),
                429,
            )

        # Extract form data
        name = request.form.get("name")
        category = request.form.get("category")
        descriptionwhy = request.form.get("descriptionwhy")
        public_benefit = request.form.get("public_benefit")
        image = request.files.get("image_file")
        geoloc = request.form.get("geoloc")
        is_global = False if geoloc else True

        if image:
            image_filename = secure_filename(image.filename)
            image_path = os.path.join(app.config["UPLOAD_FOLDER"], image_filename)
            image.save(image_path)

            current_time = datetime.now(pytz.utc) + timedelta(hours=1)

            new_project = Project(
                name=name,
                category=category,
                descriptionwhy=descriptionwhy,
                public_benefit=public_benefit,
                image_file=image_filename,
                geoloc=geoloc,
                is_global=is_global,
                author=current_user.id,
                date=current_time,
            )

            db.session.add(new_project)
            db.session.commit()

            # Update IP tracking dictionary
            submissions.append(datetime.now())
            ip_project_submissions[ip_address] = submissions

            return redirect(url_for("project_details", project_id=new_project.id))
        else:
            return redirect(url_for("submit_project"))
    metaData=g.metaData
    return render_template("neuerbeitrag/index.html", metaData=metaData)

@app.route('/robots.txt')
def robots_txt():
    return app.send_static_file('robots.txt')

@app.route('/service-worker.js')
def service_worker():
    return app.send_static_file('service-worker.js')


@app.route("/list")
@app.route("/list/pages/<int:page>")
def list_view(page=1):
    per_page = 50  # Number of projects per page
    query = Project.query.filter(Project.is_mapobject != True)

    category = request.args.get("category")
    sort = request.args.get("sort")
    search = request.args.get("search")

    # Apply category and search filters
    if category:
        query = query.filter(Project.category == category)
    if search:
        query = query.filter(Project.name.contains(search))

    # Apply sort filters
    if sort == "oldest":
        query = query.order_by(Project.date.asc())
    elif sort == "newest":
        query = query.order_by(Project.date.desc())
    elif sort == "highest":
        # Default to highest score (more upvotes)
        query = (
            query.outerjoin(Vote, Project.id == Vote.project_id)
            .group_by(Project.id)
            .order_by(func.coalesce(func.sum(Vote.upvote - Vote.downvote), 0).desc())
        )
    elif sort == "lowest":
        # Sorting by lowest score (more downvotes)
        query = (
            query.outerjoin(Vote, Project.id == Vote.project_id)
            .group_by(Project.id)
            .order_by(func.coalesce(func.sum(Vote.upvote - Vote.downvote), 0))
        )
    else:  # Change this part to default to sorting by views
        query = query.order_by(Project.view_count.desc())

    

    # Paginate the query
    paginated_projects = query.paginate(page=page, per_page=per_page, error_out=False)

    # Calculate upvote and downvote counts and percentages
    for project in paginated_projects.items:
        upvotes = sum(vote.upvote for vote in project.votes)
        downvotes = sum(vote.downvote for vote in project.votes)
        total_votes = upvotes + downvotes

        project.upvotes = upvotes
        project.downvotes = downvotes

        if total_votes > 0:
            project.upvote_percentage = upvotes / total_votes * 100
            project.downvote_percentage = downvotes / total_votes * 100
        else:
            project.upvote_percentage = 0
            project.downvote_percentage = 0
    metaData = g.metaData
    return render_template(
        "list/index.html", projects=paginated_projects.items, pagination=paginated_projects, metaData=metaData
    )


def get_project_by_id(project_id):

    return Project.query.get_or_404(project_id)


@app.route("/remove_bookmark/<int:project_id>", methods=["POST"])
def remove_bookmark(project_id):
    if not current_user.is_authenticated:
        print("User not authenticated")
        return redirect(url_for("login"))

    bookmark = Bookmark.query.filter_by(
        user_id=current_user.id, project_id=project_id
    ).first()
    if bookmark:
        db.session.delete(bookmark)
        db.session.commit()
        print(
            f"Bookmark removed for project ID {project_id} by user ID {current_user.id}"
        )
    else:
        print(
            f"No bookmark found for project ID {project_id} and user ID {current_user.id}"
        )

    return jsonify(success=True, message="Bookmark removed successfully.")


@app.route("/project_details/<int:project_id>", methods=["GET", "POST"])
def project_details(project_id):
    try:
        project = Project.query.get(project_id)
        comments = Comment.query.filter_by(project_id=project_id).all()
        is_bookmarked = (
            Bookmark.query.filter_by(
                user_id=current_user.id, project_id=project_id
            ).first()
            is not None
            if current_user.is_authenticated
            else False
        )
        is_reported = (
            Report.query.filter_by(
                user_id=current_user.id, project_id=project_id
            ).first()
            is not None
            if current_user.is_authenticated
            else False
        )

        ip_address = request.remote_addr
        WebsiteViews.add_view(ip_address)
        user_ip = request.remote_addr
        current_time = datetime.utcnow()
        last_view = ProjectView.query.filter(
            and_(
                ProjectView.project_id == project_id, ProjectView.ip_address == user_ip
            )
        ).first()

        ip_address = request.remote_addr  # Example to get IP address
        last_view = (
            ProjectView.query.filter_by(project_id=project_id, ip_address=ip_address)
            .order_by(ProjectView.last_viewed.desc())
            .first()
        )

        if last_view is None or (
            datetime.utcnow() - last_view.last_viewed > timedelta(hours=24)
        ):
            new_view = ProjectView(
                project_id=project_id,
                ip_address=ip_address,
                last_viewed=datetime.utcnow(),
            )
            db.session.add(new_view)

            # Ensure view_count is not None
            if project.view_count is None:
                project.view_count = 0
            project.view_count += 1

            db.session.commit()
            print(
                f"Project viewed by user {current_user.id if current_user.is_authenticated else 'Anonymous'} from IP {ip_address}, adding one more view. Current number of views: {project.view_count}."
            )
        else:
            print(
                f"Project viewed by user {current_user.id if current_user.is_authenticated else 'Anonymous'} from IP {ip_address}, user has however already viewed this project during the last 24 hours. Not adding a view. Current number of views: {project.view_count}."
            )

        if request.method == "POST" and current_user.is_authenticated:
            comment_text = request.form.get("comment", "").strip()
            if not (20 <= len(comment_text) <= 500):
                flash(
                    "Kommentare müssen zwischen 20 und 500 Zeichen lang sein.", "error"
                )
                return redirect(url_for("project_details", project_id=project_id))

            if not can_user_post_comment(current_user.id):
                flash(
                    "Kommentarlimit erreicht. Bitte warten Sie, bevor Sie einen weiteren Kommentar posten.",
                    "error",
                )
                return redirect(url_for("project_details", project_id=project_id))

            new_comment = Comment(
                text=comment_text, user_id=current_user.id, project_id=project_id
            )
            db.session.add(new_comment)
            db.session.commit()
            return redirect(url_for("project_details", project_id=project_id))


        

        votes = Vote.query.filter_by(project_id=project_id).all()
        upvote_count = sum(vote.upvote for vote in votes)
        downvote_count = sum(vote.downvote for vote in votes)
        total_votes = upvote_count + downvote_count
        upvote_percentage = (upvote_count / total_votes * 100) if total_votes > 0 else 0
        downvote_percentage = (
            (downvote_count / total_votes * 100) if total_votes > 0 else 0
        )

        prev_project = (
            Project.query.filter(Project.id < project_id, Project.is_mapobject == False)
            .order_by(Project.id.desc())
            .first()
        )
        next_project = (
            Project.query.filter(Project.id > project_id, Project.is_mapobject == False)
            .order_by(Project.id.asc())
            .first()
        )

        prev_project_id = prev_project.id if prev_project else None
        next_project_id = next_project.id if next_project else None

        if prev_project:
            print(f"Previous project_details page with is_mapobject=false found, it is page number {prev_project.id}")
        else:
            print("Previous project_details page with is_mapobject=false does not exist, hiding the arrowleft.")

        if next_project:
            print(f"Next project_details page with is_mapobject=false found, it is page number {next_project.id}")
        else:
            print("Next project_details page with is_mapobject=false does not exist, hiding the arrowright.")


        project_author = User.query.get(project.author)
        author_name = project_author.name if project_author else "Unknown"
        comments_with_authors = [
            {
                "text": comment.text,
                "timestamp": comment.timestamp,
                "author_name": User.query.get(comment.user_id).name
                if User.query.get(comment.user_id)
                else "Unknown",
            }
            for comment in comments
        ]
        
        is_mapobject = getattr(project, "is_mapobject", False)
        logging.debug(project)
        def remove_p_tags(text):
            pattern = re.compile(r'<p>|</p>')
            cleaned_text = pattern.sub('', text)
            return cleaned_text
        
        g.metaData['og_url']="https://stimmungskompass.ermine.at/project_details/"+str(project_id)
        g.metaData['og_description']=remove_p_tags(project.descriptionwhy)
        g.metaData['og_image']="https://stimmungskompass.ermine.at/static/usersubmissions/" + project.image_file
        g.metaData['og_title']=project.name
        metaData=g.metaData
        return render_template(
            "project_details/index.html",
            project=project,
            prev_project_id=prev_project_id,
            next_project_id=next_project_id,
            upvote_percentage=upvote_percentage,
            downvote_percentage=downvote_percentage,
            upvote_count=upvote_count,
            downvote_count=downvote_count,
            current_user=current_user,
            author_name=author_name,
            comments=comments_with_authors,
            is_mapobject=is_mapobject,
            currentUserId=current_user.id if current_user.is_authenticated else None,
            is_bookmarked=is_bookmarked,
            is_reported=is_reported,
            metaData=metaData
        )
    except Exception as e:
        app.logger.error("Error in project_details route: %s", str(e))
        return str(e)  # Or redirect to a generic error page


@app.route("/report/<int:project_id>", methods=["POST"])
def toggle_report(project_id):
    if not current_user.is_authenticated:
        return jsonify({"success": False, "message": "User not authenticated"}), 403

    try:
        report = Report.query.filter_by(
            user_id=current_user.id, project_id=project_id
        ).first()
        if report:
            db.session.delete(report)
            db.session.commit()
            return (
                jsonify(
                    {"success": True, "message": "Report removed", "reported": False}
                ),
                200,
            )
        else:
            new_report = Report(user_id=current_user.id, project_id=project_id)
            db.session.add(new_report)
            db.session.commit()
            return (
                jsonify(
                    {"success": True, "message": "Project reported", "reported": True}
                ),
                200,
            )
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": "An error occurred"}), 500


@app.route("/bookmark/<int:project_id>", methods=["POST"])
def toggle_bookmark(project_id):
    if not current_user.is_authenticated:
        return jsonify({"success": False, "message": "User not authenticated"}), 403

    try:
        bookmark = Bookmark.query.filter_by(
            user_id=current_user.id, project_id=project_id
        ).first()
        if bookmark:
            db.session.delete(bookmark)
            db.session.commit()
            return (
                jsonify(
                    {
                        "success": True,
                        "message": "Bookmark removed",
                        "bookmarked": False,
                    }
                ),
                200,
            )
        else:
            new_bookmark = Bookmark(user_id=current_user.id, project_id=project_id)
            db.session.add(new_bookmark)
            db.session.commit()
            return (
                jsonify(
                    {
                        "success": True,
                        "message": "Project bookmarked",
                        "bookmarked": True,
                    }
                ),
                200,
            )
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": "An error occurred"}), 500


@app.route("/check_comment_limit")
def check_comment_limit():
    if not current_user.is_authenticated:
        return jsonify(limit_reached=False, current_count=0)

    time_limit = datetime.now() - timedelta(minutes=15)
    recent_comments = (
        Comment.query.filter(
            Comment.user_id == current_user.id, Comment.timestamp > time_limit
        )
        .order_by(Comment.timestamp.asc())
        .all()
    )
    recent_comments_count = len(recent_comments)

    if recent_comments_count >= 5:
        oldest_comment_time = recent_comments[0].timestamp
        reset_time = oldest_comment_time + timedelta(minutes=15)
        return jsonify(
            limit_reached=True,
            reset_time=reset_time.isoformat(),
            current_count=recent_comments_count,
        )
    else:
        return jsonify(
            limit_reached=False, reset_time=None, current_count=recent_comments_count
        )


@app.route("/preview_delete_map_objects_by_date", methods=["POST"])
@login_required
def preview_delete_map_objects_by_date():
    if current_user.id != 1:
        return jsonify(success=False, error="Unauthorized access.")

    data = request.get_json()
    from_date = datetime.strptime(data["fromDate"], "%Y-%m-%d")
    to_date = datetime.strptime(data["toDate"], "%Y-%m-%d")

    map_objects = Project.query.filter(
        Project.is_mapobject == True, Project.date >= from_date, Project.date <= to_date
    ).all()
    map_objects_info = [
        {"id": obj.id, "descriptionwhy": obj.descriptionwhy} for obj in map_objects
    ]

    return jsonify(success=True, mapObjects=map_objects_info)


@app.route("/delete_map_objects_by_date", methods=["POST"])
@login_required
def delete_map_objects_by_date():
    if current_user.id != 1:  # Ensure it's the admin
        return jsonify(success=False, error="Unauthorized access.")

    data = request.get_json()
    from_date = datetime.strptime(data["fromDate"], "%Y-%m-%d")
    to_date = datetime.strptime(data["toDate"], "%Y-%m-%d")

    map_objects_to_delete = Project.query.filter(
        Project.is_mapobject == True, Project.date >= from_date, Project.date <= to_date
    )
    deleted_count = map_objects_to_delete.delete()

    db.session.commit()
    app.logger.debug(
        f"Deleted {deleted_count} map objects between {from_date} and {to_date}"
    )

    return jsonify(
        success=True, message=f"{deleted_count} map objects deleted successfully."
    )


@app.route("/admintools", methods=["GET", "POST"])
@login_required
def admintools():
    # Check if the user is the admin
    
    if current_user.id != 1:
        flash("", "danger")
        return redirect(url_for("index"))
    
    # Check for OTP verification
    if "admin_verified" not in session or not session["admin_verified"]:
        # Generate OTP
        otp = randint(100000, 999999)
        client = Client(account_sid, auth_token)
        message = client.messages.create(
            body=f"Stimmungskompass: Um sich bei Admintools anzumelden, verwenden Sie OTP: {otp}",
            from_=twilio_number,
            to=current_user.phone_number,
        )
        session["admin_otp"] = otp
        return redirect(url_for("verify_admin_otp"))

    # Admin tools logic begins here
    users = User.query.all()
    for user in users:
        user.project_count = Project.query.filter_by(
            author=user.id, is_mapobject=False
        ).count()
        user.map_object_count = Project.query.filter_by(
            author=user.id, is_mapobject=True
        ).count()
        user.comment_count = Comment.query.filter_by(user_id=user.id).count()

    comments = Comment.query.all()
    for comment in comments:
        project = Project.query.get(comment.project_id)
        user = User.query.get(comment.user_id)
        comment.project_name = project.name if project else "Unknown Project"
        comment.author_name = user.name if user else "Unknown Author"
        comment.author_id = user.id if user else "Unknown ID"

    top_viewed_projects = (
        Project.query.order_by(Project.view_count.desc()).limit(5).all()
    )
    top_rated_projects = (
        db.session.query(
            Project.id, Project.name, func.count(Vote.id).label("upvote_count")
        )
        .join(Vote, Project.id == Vote.project_id)
        .filter(Vote.upvote == True)
        .group_by(Project.id)
        .order_by(desc("upvote_count"))
        .limit(5)
        .all()
    )

    top_commented_projects_query = (
        db.session.query(
            Project.id, Project.name, func.count(Comment.id).label("comments_count")
        )
        .join(Comment, Project.id == Comment.project_id)
        .group_by(Project.id)
        .order_by(func.count(Comment.id).desc())
        .limit(5)
        .all()
    )

    top_commented_projects = [
        {"id": project_id, "name": project_name, "comments_count": comments_count}
        for project_id, project_name, comments_count in top_commented_projects_query
    ]

    # Top categories
    category_counts = Counter(
        [project.category for project in Project.query.all()]
    ).most_common(5)

    # Top active accounts
    active_users = (
        User.query.outerjoin(Project, User.id == Project.author)
        .group_by(User.id)
        .order_by(func.count(Project.id).desc())
        .limit(5)
        .all()
    )

    app.logger.debug("Top statistics calculated for admin tools")

    # POST request handling
    if request.method == "POST":
        project_id = request.form.get("project_id")

        if "unmark_important" in request.form:
            project = Project.query.get(project_id)
            if project:
                project.is_important = False
                db.session.commit()
                flash("Project unmarked as important", "success")
            else:
                flash("Project not found for unmarking as important", "error")

        elif "unmark_featured" in request.form:
            project = Project.query.get(project_id)
            if project:
                project.is_featured = False
                db.session.commit()
                flash("Project unmarked as featured", "success")
            else:
                flash("Project not found for unmarking as featured", "error")

        elif "remove_reports" in request.form:
            project_id = request.form.get("project_id")
            project = Project.query.get(project_id)

            if project:
                # Remove all reports associated with the project
                Report.query.filter_by(project_id=project_id).delete()

                # Commit the changes to the database
                db.session.commit()

                flash(
                    "Reports have been removed from the project successfully.",
                    "success",
                )
            else:
                flash("Project not found for removing reports.", "error")

        elif "delete_project" in request.form:
            project = Project.query.get(project_id)
            if project:
                db.session.delete(project)
                db.session.commit()
                flash(f"Project {project_id} successfully deleted.", "success")
            else:
                flash(f"Project {project_id} not found for deletion.", "error")

        return redirect(url_for("admintools"))

    # GET request logic with pagination and sorting
    sort = request.args.get("sort", "score_desc")
    search_query = request.args.get("search", "")
    page = request.args.get("page", 1, type=int)
    per_page = 50

    map_object_page = request.args.get("map_object_page", 1, type=int)
    map_object_per_page = 50  # Define the number of Notizens per page

    comment_page = request.args.get("comment_page", 1, type=int)
    comment_per_page = 50  # Adjust the number of comments per page as needed

    user_page = request.args.get("user_page", 1, type=int)
    user_per_page = 50  # Define the number of users per page

    search_user_id = request.args.get("searchUserById", type=int)
    search_user_name = request.args.get("searchUserByName", "")
    search_comment_query = request.args.get("searchComment", "")
    search_map_object_query = request.args.get("searchMapObject", "")

    categories = db.session.query(Project.category).distinct().all()
    categories = [
        category[0] for category in db.session.query(Project.category.distinct()).all()
    ]

    app.logger.debug(f"Found categories: {categories}")

    # Query Notizens with the search filter
    map_object_query = Project.query.filter_by(is_mapobject=True)
    if search_map_object_query:
        map_object_query = map_object_query.filter(
            Project.descriptionwhy.ilike(f"%{search_map_object_query}%")
        )

    search_comment_by_user_id = request.args.get("searchCommentByUserId")
    search_comment_query = request.args.get("searchComment")
    comment_query = Comment.query

    if search_comment_by_user_id:
        comment_query = comment_query.filter(
            Comment.user_id == int(search_comment_by_user_id)
        )
    if search_comment_query:
        comment_query = comment_query.filter(
            Comment.text.ilike(f"%{search_comment_query}%")
        )

    comment_sort = request.args.get(
        "comment_sort", "newest"
    )  # Default sorting is by newest

    # Sorting logic
    if comment_sort == "oldest":
        comment_query = comment_query.order_by(Comment.timestamp)
    elif comment_sort == "newest":
        comment_query = comment_query.order_by(Comment.timestamp.desc())

    query = Project.query.filter(
        Project.is_mapobject == False
    )  # Filtern for non-mapobject projects

    if search_query:
        query = query.filter(Project.name.contains(search_query))
    # User query with optional search
    user_query = User.query
    if search_user_id:
        user_query = user_query.filter(User.id == search_user_id)
    if search_user_name:
        user_query = user_query.filter(User.name.ilike(f"%{search_user_name}%"))

    # Adjust the sorting logic
    if sort == "comments":
        comments_subquery = (
            db.session.query(
                Comment.project_id, func.count("*").label("comments_count")
            )
            .group_by(Comment.project_id)
            .subquery()
        )
        query = query.outerjoin(
            comments_subquery, Project.id == comments_subquery.c.project_id
        ).order_by(desc(comments_subquery.c.comments_count))

    elif sort == "oldest":
        query = query.order_by(Project.date)
    elif sort == "newest":
        query = query.order_by(desc(Project.date))
    elif sort == "category":
        query = query.order_by(Project.category)
    elif sort == "user_id":
        query = query.order_by(Project.author)
    elif sort == "upvotes":
        query = (
            query.outerjoin(Vote, Project.id == Vote.project_id)
            .group_by(Project.id)
            .order_by(func.sum(Vote.upvote).desc())
        )
    elif sort == "downvotes":
        query = (
            query.outerjoin(Vote, Project.id == Vote.project_id)
            .group_by(Project.id)
            .order_by(func.sum(Vote.downvote).desc())
        )
    elif sort == "alpha_asc":
        query = query.order_by(asc(Project.name))
    elif sort == "alpha_desc":
        query = query.order_by(desc(Project.name))
    else:
        query = (
            query.outerjoin(Vote, Project.id == Vote.project_id)
            .group_by(Project.id)
            .order_by(func.sum(Vote.upvote - Vote.downvote).desc())
        )

    paginated_users = user_query.paginate(
        page=user_page, per_page=user_per_page, error_out=False
    )
    paginated_comments = comment_query.paginate(
        page=comment_page, per_page=comment_per_page, error_out=False
    )
    paginated_map_objects = map_object_query.paginate(
        page=map_object_page, per_page=map_object_per_page, error_out=False
    )
    paginated_projects = query.paginate(page=page, per_page=per_page, error_out=False)
    print("Total items:", paginated_projects.total)
    
    metaData=g.metaData
    print("Total pages:", paginated_projects.pages)

    # Calculate upvotes and downvotes for each project
    for project in paginated_projects.items:
        upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
        downvotes = Vote.query.filter_by(project_id=project.id, downvote=True).count()

        project.upvotes = upvotes
        project.downvotes = downvotes

        total_votes = upvotes + downvotes
        if total_votes > 0:
            project.upvote_percentage = (upvotes / total_votes) * 100
            project.downvote_percentage = (downvotes / total_votes) * 100
        else:
            project.upvote_percentage = 0
            project.downvote_percentage = 0

    # Updating user statistics for all users
    users = User.query.all()
    for user in users:
        user.project_count = Project.query.filter_by(
            author=user.id, is_mapobject=False
        ).count()
        user.map_object_count = Project.query.filter_by(
            author=user.id, is_mapobject=True
        ).count()
        user.comment_count = Comment.query.filter_by(user_id=user.id).count()

    # Updating comment information
    comments = Comment.query.all()
    for comment in comments:
        project = Project.query.get(comment.project_id)
        user = User.query.get(comment.user_id)
        comment.project_name = project.name if project else "Unknown Project"
        comment.author_name = user.name if user else "Unknown Author"
        comment.author_id = user.id if user else "Unknown ID"

    # Prepare lists of important and featured projects
    important_projects = [
        project for project in paginated_projects.items if project.is_important
    ]
    featured_projects = [
        project for project in paginated_projects.items if project.is_featured
    ]
    reported_projects = [
        project
        for project in paginated_projects.items
        if Report.query.filter_by(project_id=project.id).first()
    ]
    user_count = User.query.count()
    comment_count = Comment.query.count()
    project_count = Project.query.filter_by(is_mapobject=False).count()
    mapobject_count = Project.query.filter_by(is_mapobject=True).count()
    bookmark_count = Bookmark.query.count()
    
    # Check if it's an AJAX request
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        request_type = request.args.get("request_type")
        if request_type == "map_object":
            return render_template(
                "partials/mapobject_list_section.html",
                paginated_map_objects=paginated_map_objects, metaData=metaData,
            )
        elif request_type == "project":
            return render_template(
                "partials/project_list_section.html",
                paginated_projects=paginated_projects, 
                sort=sort,
                search_query=search_query, metaData=metaData,
            )
        elif request_type == "comment":
            return render_template(
                "partials/comments_section.html", paginated_comments=paginated_comments,  metaData=metaData
            ) 
        elif request_type == "user":
            return render_template(
                "partials/user_list_section.html", paginated_users=paginated_users, metaData=metaData,
            )
    # Normal request
    return render_template(
        "admintools.html",
        paginated_projects=paginated_projects,
        paginated_map_objects=paginated_map_objects,
        paginated_comments=paginated_comments,
        paginated_users=paginated_users,
        sort=sort,
        categories=categories,
        search_query=search_query,
        user_count=user_count,
        comment_count=comment_count,
        project_count=project_count,
        mapobject_count=mapobject_count,
        bookmark_count=bookmark_count,
        users=users,
        important_projects=important_projects,
        featured_projects=featured_projects,
        reported_projects=reported_projects,
        top_viewed_projects=top_viewed_projects,
        top_rated_projects=top_rated_projects,
        top_commented_projects=top_commented_projects,
        category_counts=category_counts,
        active_users=active_users,
        # metaData=metaData,
    )


@app.route('/unmark_important/<int:projectId>', methods=['POST'])
@login_required
def unmark_important(projectId):
    project = Project.query.get_or_404(projectId)
    project.is_important = False
    db.session.commit()
    return jsonify({'success': True, 'message': 'Project unmarked as important.', 'projectId': projectId})


@app.route('/unmark_featured/<int:projectId>', methods=['POST'])
def unmark_featured(projectId):
    # Example logic to unmark a project as featured
    # Replace this with your actual database update logic
    project = Project.query.get_or_404(projectId)
    project.is_featured = False
    db.session.commit()

    return jsonify({'success': True, 'message': 'Projekt erfolgreich als nicht ausgewählt markiert.'})
    
    

@app.route('/mark_important/<int:projectId>', methods=['POST'])
@login_required
def mark_important(projectId):
    project = Project.query.get_or_404(projectId)
    project.is_important = True
    db.session.commit()
    return jsonify({'success': True, 'message': 'Project marked as important.', 'projectId': projectId})



@app.route('/mark_featured/<int:project_id>', methods=['POST'])
@login_required
def mark_featured(project_id):
    project = Project.query.get_or_404(project_id)

    # Calculate upvotes, downvotes, and their percentages
    upvotes = sum(1 for vote in project.votes if vote.upvote)
    downvotes = sum(1 for vote in project.votes if not vote.upvote)  # Assuming 'not vote.upvote' marks a downvote
    total_votes = upvotes + downvotes
    upvote_percentage = (upvotes / total_votes * 100) if total_votes > 0 else 0
    downvote_percentage = (downvotes / total_votes * 100) if total_votes > 0 else 0

    # Mark the project as featured
    project.is_featured = True
    db.session.commit()  # Save the changes to the database

    # Construct the response data
    response_data = {
        "success": True,
        "project_id": project.id,
        "is_featured": project.is_featured,
        "project_details": {
            "name": project.name,
            "date": project.date.strftime('%Y-%m-%d'),  # Adjust the date format as per your requirements
            "view_count": project.view_count,
            "image_file": project.image_file,
            "descriptionwhy": project.descriptionwhy,
            "upvotes": upvotes,
            "downvotes": downvotes,
            "upvote_percentage": upvote_percentage,
            "downvote_percentage": downvote_percentage
        }
    }

    # Return the success response with detailed project information
    return jsonify(response_data)
    
    
@app.route('/delete_map_object/<int:map_object_id>', methods=['POST'])
@login_required
def delete_map_object(map_object_id):
    # Authentication and authorization checks here
    map_object = MapObject.query.get_or_404(map_object_id)
    db.session.delete(map_object)
    db.session.commit()
    return jsonify(success=True)



@app.route("/verify_admin_otp", methods=["GET", "POST"])
@login_required
def verify_admin_otp():
    # Ensure the user is an admin
    if current_user.id != 1:
        flash("Access Denied: You are not authorized to perform this action.", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        entered_otp = request.form.get("otp")

        # Passwort bestätigen
        if "admin_otp" in session and str(session["admin_otp"]) == entered_otp:
            session["admin_verified"] = True
            flash("OTP Verified. Access granted to admin tools.", "success")
            return redirect(url_for("admintools"))
        else:
            flash("Invalid OTP. Please try again.", "danger")
    metaData=g.metaData
    return render_template(
        "verify_admin_otp/index.html", metaData=metaData,
    )  # Ensure this template exists for OTP input


import os
import shutil


@app.route("/delete_my_data", methods=["POST"])
@login_required
def delete_my_data():
    try:
        user_id = current_user.id

        # Löschen user's votes
        Vote.query.filter_by(user_id=user_id).delete()

        # Löschen user's comments
        Comment.query.filter_by(user_id=user_id).delete()

        # Löschen user's projects and associated files
        projects = Project.query.filter_by(author=user_id).all()
        for project in projects:
            total_votes = sum(vote.upvote + vote.downvote for vote in project.votes)
            project.upvotes = sum(vote.upvote for vote in project.votes)
            project.downvotes = sum(vote.downvote for vote in project.votes)

            if total_votes > 0:
                project.upvote_percentage = project.upvotes / total_votes * 100
                project.downvote_percentage = project.downvotes / total_votes * 100
            else:
                project.upvote_percentage = 0
                project.downvote_percentage = 0

            # Löschen associated files (if applicable)
            project_files_path = os.path.join(
                "actual/path/to/project_files", str(project.id)
            )
            if os.path.exists(project_files_path):
                shutil.rmtree(project_files_path)

            # Löschen the project record from the database
            db.session.delete(project)

        # Löschen user account
        user = User.query.filter_by(id=user_id).first()
        if user:
            db.session.delete(user)

        # Commit changes to the database
        db.session.commit()

        # Log out the user
        logout_user()

        # Return JSON response
        return jsonify(
            {"success": True, "message": "Your data has been deleted successfully."}
        )
    except Exception as e:
        logging.error(f"Error in delete_my_data: {e}")
        return (
            jsonify(
                {
                    "success": False,
                    "message": "An error occurred while deleting your data.",
                }
            ),
            500,
        )


@app.route("/upload_screenshot", methods=["POST"])
def upload_screenshot():
    screenshot = request.files["screenshot"]
    filename = "screenshot.png"
    filepath = os.path.join("static/screenshots", filename)
    screenshot.save(filepath)

    # Generate a shareable link
    link = request.host_url + filepath
    return jsonify({"link": link})


@app.route("/delete_user", methods=["POST"])
@login_required
def delete_user():
    try:
        user_id_to_delete = request.form.get("user_id")
        user_to_delete = User.query.get(user_id_to_delete)
        if user_to_delete:
            # Löschen user's votes
            Vote.query.filter_by(user_id=user_id_to_delete).delete()

            # Löschen user's comments
            Comment.query.filter_by(user_id=user_id_to_delete).delete()

            # Löschen user's projects and associated files
            projects_to_delete = Project.query.filter_by(author=user_id_to_delete).all()
            for project in projects_to_delete:
                # Löschen associated votes and comments for each project
                Vote.query.filter_by(project_id=project.id).delete()
                Comment.query.filter_by(project_id=project.id).delete()

                # Löschen project files (you may need to implement this logic)
                # For example, if you store project images in a folder, you can delete them here

                # Finally, delete the project itself
                db.session.delete(project)

            # Löschen user account
            db.session.delete(user_to_delete)

            # Commit changes to the database
            db.session.commit()

            return jsonify(
                {
                    "success": True,
                    "message": "User and their contributions have been deleted successfully.",
                }
            )
        else:
            return (
                jsonify({"success": False, "message": "User not found for deletion."}),
                404,
            )
    except Exception as e:
        logging.error(f"Error in delete_user: {e}")
        return (
            jsonify(
                {
                    "success": False,
                    "message": "An error occurred while deleting the user and contributions.",
                }
            ),
            500,
        )


@app.route("/downvote/<int:project_id>", methods=["POST"])
@login_required
def downvote(project_id):
    project = Project.query.get_or_404(project_id)
    existing_downvote = Downvote.query.filter_by(
        user_id=current_user.id, project_id=project.id
    ).first()

    if existing_downvote:
        flash('You have already downvoted this project.', 'info')
        return redirect(url_for("list_view"))

    downvote = Downvote(
        user_id=current_user.id, project_id=project.id, ip_address=request.remote_addr
    )
    db.session.add(downvote)
    db.session.commit()
    flash('Your downvote has been recorded!', 'success')
    return redirect(url_for("list_view"))


@app.route("/vote/<int:project_id>/<string:vote_type>", methods=["POST"])
@login_required
def vote(project_id, vote_type):
    project = Project.query.get_or_404(project_id)

    # Check for existing votes by the user for this project
    existing_vote = Vote.query.filter_by(
        user_id=current_user.id, project_id=project.id
    ).first()

    if existing_vote:
        db.session.delete(existing_vote)

    new_vote = Vote(user_id=current_user.id, project_id=project.id)
    if vote_type == "upvote":
        new_vote.upvote = True
        new_vote.downvote = False
        app.logger.debug(
            f"Upvote received for project {project_id} by user {current_user.id}"
        )
    elif vote_type == "downvote":
        new_vote.downvote = True
        new_vote.upvote = False
        app.logger.debug(
            f"Downvote received for project {project_id} by user {current_user.id}"
        )

    db.session.add(new_vote)
    db.session.commit()

    # Re-fetch the project to get updated vote counts
    project = Project.query.get_or_404(project_id)
    upvote_count = sum(vote.upvote for vote in project.votes)
    downvote_count = sum(vote.downvote for vote in project.votes)
    total_votes = upvote_count + downvote_count
    upvote_percentage = (upvote_count / total_votes * 100) if total_votes > 0 else 0
    downvote_percentage = (downvote_count / total_votes * 100) if total_votes > 0 else 0

    # Return updated vote data
    return jsonify(
        {
            "success": True,
            "message": "Vote recorded",
            "upvote_count": upvote_count,
            "downvote_count": downvote_count,
            "upvote_percentage": upvote_percentage,
            "downvote_percentage": downvote_percentage,
        }
    )


@app.route("/comment/<int:project_id>", methods=["POST"])
@login_required
def comment(project_id):
    # Check if the user can post a comment
    if not can_user_post_comment(current_user.id):
        # If the user has exceeded the comment limit, return an appropriate response
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Kommentarlimit erreicht. Bitte warten Sie, bevor Sie einen weiteren Kommentar veröffentlichen.",
                }
            ),
            429,
        )  # 429 Too Many Requests

    project = Project.query.get_or_404(project_id)
    comment_text = request.form.get("comment")

    # Add one hour to the current timestamp
    timestamp = datetime.now() + timedelta(hours=0)

    new_comment = Comment(
        text=comment_text,
        user_id=current_user.id,
        project_id=project_id,
        timestamp=timestamp,
    )
    db.session.add(new_comment)
    db.session.commit()

    return jsonify(
        {
            "success": True,
            "text": new_comment.text,
            "author_name": f"{current_user.name}",
            "timestamp": new_comment.timestamp.strftime("%d.%m.%Y %H:%M"),
        }
    )


@app.route("/bookmarked")
def bookmarked():
    metaData=g.metaData
    if current_user.is_authenticated:
        user_bookmarked_projects = (
            Project.query.join(Bookmark, Bookmark.project_id == Project.id)
            .filter(Bookmark.user_id == current_user.id)
            .all()
        )
        return render_template(
            "bookmarked.html", user_bookmarked_projects=user_bookmarked_projects, metaData=metaData
        )
    else:
        return redirect(url_for("login"))


@app.route("/profil")
@app.route(
    "/profil/projects/<int:project_page>/map_objects/<int:map_object_page>/comments/<int:comment_page>"
)
def profil(project_page=1, map_object_page=1, comment_page=1):
    per_page = 50  # Number of items per page
    user_statistics = {}
    paginated_projects = []
    paginated_map_objects = []
    paginated_comments = []
    bookmarks = []  # Default value for bookmarks
    bookmarked_projects = []  # Default value for bookmarked projects

    ip_address = request.remote_addr
    WebsiteViews.add_view(ip_address)

    if current_user.is_authenticated:
        bookmarks = Bookmark.query.filter_by(user_id=current_user.id).all()
        paginated_projects = (
            Project.query.filter_by(author=current_user.id, is_mapobject=False)
            .order_by(Project.date.desc())
            .paginate(page=project_page, per_page=per_page, error_out=False)
        )

        total_map_objects = Project.query.filter_by(
            author=current_user.id, is_mapobject=True
        ).count()
        max_page = (
            total_map_objects + per_page - 1
        ) // per_page  # Calculate the max page number
        map_object_page = min(map_object_page, max_page)  # Adjust the current page

        paginated_map_objects = (
            Project.query.filter_by(author=current_user.id, is_mapobject=True)
            .order_by(Project.date.desc())
            .paginate(page=map_object_page, per_page=per_page, error_out=False)
        )

        paginated_comments = (
            db.session.query(Comment, Project.name)
            .join(Project, Comment.project_id == Project.id)
            .filter(Comment.user_id == current_user.id)
            .order_by(Comment.timestamp.desc())
            .paginate(page=comment_page, per_page=per_page, error_out=False)
        )
        logging.debug(paginated_comments)
        bookmarked_projects = (
            Project.query.join(Bookmark, Bookmark.project_id == Project.id)
            .filter(Bookmark.user_id == current_user.id)
            .all()
        )

        # Count Notizens separately
        map_objects_count = Project.query.filter_by(
            author=current_user.id, is_mapobject=True
        ).count()

        for project in paginated_projects.items:
            upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
            downvotes = Vote.query.filter_by(
                project_id=project.id, downvote=True
            ).count()
            total_votes = upvotes + downvotes
            project.upvotes = upvotes
            project.downvotes = downvotes
            project.upvote_percentage = (
                (upvotes / total_votes * 100) if total_votes > 0 else 0
            )
            project.downvote_percentage = (
                (downvotes / total_votes * 100) if total_votes > 0 else 0
            )

        for project in bookmarked_projects:
            upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
            downvotes = Vote.query.filter_by(
                project_id=project.id, downvote=True
            ).count()
            total_votes = upvotes + downvotes
            project.upvotes = upvotes
            project.downvotes = downvotes
            project.upvote_percentage = (
                (upvotes / total_votes * 100) if total_votes > 0 else 0
            )
            project.downvote_percentage = (
                (downvotes / total_votes * 100) if total_votes > 0 else 0
            )

        # Prepare user statistics
        num_projects = Project.query.filter_by(
            author=current_user.id, is_mapobject=False
        ).count()
        num_map_objects = Project.query.filter_by(
            author=current_user.id, is_mapobject=True
        ).count()
        num_comments = Comment.query.filter_by(user_id=current_user.id).count()
        bookmarked_projects = (
            Project.query.join(Bookmark, Bookmark.project_id == Project.id)
            .filter(Bookmark.user_id == current_user.id)
            .all()
        )
        most_viewed_project = (
            Project.query.filter_by(author=current_user.id)
            .order_by(Project.view_count.desc())
            .first()
        )

        # Find the most successful project
        most_successful_project = None
        max_upvotes = 0
        for project in paginated_projects.items:
            upvotes = Vote.query.filter_by(project_id=project.id, upvote=True).count()
            if upvotes > max_upvotes:
                max_upvotes = upvotes
                most_successful_project = project

        user_statistics = {
            "num_projects": num_projects,
            "num_map_objects": num_map_objects,
            "num_comments": num_comments,
            "most_successful_project": most_successful_project,
            "most_viewed_project": most_viewed_project,
        }
    else:
        paginated_projects = None
        paginated_map_objects = None
        paginated_comments = None
    metaData=g.metaData
    # Check if it's an AJAX request
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        requested_section = request.args.get("section")

        if requested_section == "comments":
            return render_template(
                "partials/comments_section_profil.html",
                comment_pagination=paginated_comments,
                project_page=project_page,
                map_object_page=map_object_page,
                user_statistics=user_statistics,
                metaData=metaData,
            )
        elif requested_section == "map_objects":
            return render_template(
                "partials/map_objects_section.html",
                map_object_pagination=paginated_map_objects,
                project_page=project_page,
                comment_page=comment_page,
                user_statistics=user_statistics,
                metaData=metaData
            )
        elif requested_section == "projects":
            return render_template(
                "partials/projects_section.html",
                project_pagination=paginated_projects,
                map_object_page=map_object_page,
                comment_page=comment_page,
                user_statistics=user_statistics,
                metaData=metaData
            )

    # Render the full page for a normal request
    return render_template(
        "profil/index.html",
        project_pagination=paginated_projects,
        map_object_pagination=paginated_map_objects,
        user_bookmarked_projects=bookmarked_projects,
        comment_pagination=paginated_comments,
        user_statistics=user_statistics,
        is_authenticated=current_user.is_authenticated,
        bookmarks=bookmarks,
        metaData=metaData,
    )


@app.route("/get_user_statistics")
@login_required
def get_user_statistics():
    try:
        num_map_objects = Project.query.filter_by(
            author=current_user.id, is_mapobject=True
        ).count()
        return jsonify({"num_map_objects": num_map_objects})
    except Exception as e:
        app.logger.error(f"Error fetching user statistics: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/erfolge")
def erfolge():
    metaData=g.metaData
    # Additional logic can be added here if needed
    return render_template("erfolge.html", metaData=metaData)


@app.route("/ueber")
def ueber():
    ip_address = request.remote_addr
    metaData=g.metaData
    WebsiteViews.add_view(ip_address)
    # Additional logic can be added here if needed
    return render_template("ueber/index.html", metaData=metaData)

@app.route("/privacy")
def privacy():
    ip_address = request.remote_addr
    WebsiteViews.add_view(ip_address)
    metaData=g.metaData
    # Additional logic can be added here if needed
    return render_template("privacy/index.html", metaData=metaData)


@app.route("/delete_project/<int:project_id>", methods=["POST"])
@login_required
def delete_project(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        app.logger.debug(
            f"User {current_user.id} attempting to delete project {project_id}, owned by user {project.author}"
        )

        # Convert project.author to int for comparison
        project_author_id = int(project.author)
        app.logger.debug(f"Converted project.author to int: {project_author_id}")

        if project_author_id == current_user.id:
            db.session.delete(project)
            db.session.commit()
            app.logger.debug(f"Project {project_id} deleted successfully.")
            return jsonify(
                {"success": True, "message": "Project deleted successfully."}
            )  # Return JSON response
        else:
            app.logger.debug(f"Permission denied to delete project {project_id}.")
            return (
                jsonify({"success": False, "message": "Permission denied."}),
                403,
            )  # Return JSON response for forbidden access
    except Exception as e:
        app.logger.error(f"Error deleting project: {e}")
        return (
            jsonify({"success": False, "message": str(e)}),
            500,
        )  # Return JSON response for internal server error


@app.route("/download_my_data")
@login_required
def download_my_data():
    user_id = current_user.id

    # Fetch user data
    user_data = User.query.filter_by(id=user_id).first()

    # Fetch user's projects
    projects = Project.query.filter_by(author=user_id).all()
    projects_data = [project.to_dict() for project in projects]

    # Fetch user's comments
    comments = Comment.query.filter_by(user_id=user_id).all()
    comments_data = [comment.to_dict() for comment in comments]

    # Fetch user's votes
    votes = Vote.query.filter_by(user_id=user_id).all()
    votes_data = [vote.to_dict() for vote in votes]

    # Aggregate data
    data = {
        "user_info": {
            "name": user_data.name,
            "phone_number": user_data.phone_number,
            "account_creation": user_data.account_creation.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "ip_address": user_data.ip_address,
        },
        "projects": projects_data,
        "comments": comments_data,
        "votes": votes_data,
    }

    # Convert data to JSON format
    response = jsonify(data)
    response.headers[
        "Content-Disposition"
    ] = f"attachment; filename=user_{user_id}_data.json"
    return response



@app.route("/delete_comment/<int:comment_id>", methods=["POST"])
@login_required
def delete_comment(comment_id):
    comment = Comment.query.get_or_404(comment_id)
    if comment.user_id == current_user.id or current_user.id == 1:  # Assuming admin has user ID 1
        db.session.delete(comment)
        db.session.commit()
        return jsonify(success=True, comment_id=comment_id)
    else:
        return jsonify(success=False, error="Permission denied"), 403


@app.route("/verify_otp", methods=["GET", "POST"])
def verify_otp():
    if request.method == "POST":
        user_otp = request.form.get("otp")
        if "user_data" in session and session["user_data"]["otp"] == int(user_otp):
            # Handle New User Registration
            user_data = session.pop("user_data")

            new_user = User(
                phone_number=user_data["phone_number"],
                password_hash=generate_password_hash(user_data["password"]),
                name=user_data["name"],
                ip_address=user_data["ip_address"],
            )
            db.session.add(new_user)
            db.session.commit()

            login_user(new_user)
            logging.debug("New user registered: %s", new_user.phone_number)
            return jsonify({"success": True, "message": "Konto erfolgreich erstellt"})

        elif "reset_otp" in session and session["reset_otp"] == int(user_otp):
            # Handle Password Reset
            phone_number = session.pop("phone_number")
            user = User.query.filter_by(phone_number=phone_number).first()
            if user:
                new_password = request.form.get("new_password")
                user.set_password(new_password)
                db.session.commit()
                logging.debug(
                    f"Password reset for user with Ihre Handynummer {phone_number}"
                )
                flash("Your password has been reset successfully.", "success")
                return redirect(url_for("login"))
            else:
                flash("User not found.", "error")

        else:
            logging.debug("OTP verification failed")
            flash("Invalid OTP", "error")
    metaData=g.metaData
    return render_template("verify_otp/index.html", metaData=metaData)


@app.route("/password_recovery", methods=["GET", "POST"])
def password_recovery():
    if request.method == "POST":
        phone_number = request.form["phone_number"]
        user = User.query.filter_by(phone_number=phone_number).first()
        if user:
            otp = randint(100000, 999999)
            client = Client(account_sid, auth_token)
            client.messages.create(
                body=f"Stimmungskompass: Ihr OTP ist: {otp}",
                from_=twilio_number,
                to=phone_number,
            )
            session["phone_number"] = phone_number
            session["otp"] = otp
            return redirect(url_for("verify_otp"))
        else:
            flash("Ihre Handynummer not found", "error")
    metaData=g.metaData
    return render_template("password_recovery.html", metaData=metaData)


@app.route("/download_data")
def download_data():
    # Querying data from the database
    projects = Project.query.all()
    votes = Vote.query.all()
    comments = Comment.query.all()

    # Convert data to JSON format
    data = {
        "projects": [project.to_dict() for project in projects],
        "votes": [vote.to_dict() for vote in votes],
        "comments": [comment.to_dict() for comment in comments],
    }

    # Create a zip file of user submissions
    zip_path = zip_user_submissions()  # Ensure this function is defined

    if zip_path and zip_path.is_file():
        data["image_zip_file"] = url_for("static", filename="usersubmissions.zip")
    else:
        data["image_zip_file"] = "No images available to download"

    # Create a formatted JSON response
    response = Response(
        json.dumps(data, default=str, indent=4),  # Adding indentation
        mimetype="application/json",
        headers={"Content-Disposition": "attachment;filename=data.json"},
    )

    return response


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route("/login", methods=["GET", "POST"])
def login():

    # Capture 'next' parameter or set to index if not present
    next_page = request.args.get("next") or url_for("index")
    logging.debug(
        f"'next' parameter detected, user will be redirected to {next_page} after logging in."
    )

    if request.method == "POST":
        username_or_phone = request.form.get("username_or_phone")
        password = request.form.get("password")
        user = User.query.filter_by(phone_number=username_or_phone).first()

        if user and user.check_password(password):
            login_user(user)
            logging.debug("Login successful")
            return jsonify(success=True, next=next_page)
        else:
            logging.debug("Login failed")
            return jsonify(success=False)
    metaData=g.metaData
    return render_template("login/index.html", next=next_page, metaData=metaData)


@app.route("/vote/<int:project_id>", methods=["GET", "POST"])
@login_required
def single_vote(project_id):
    project = Project.query.get_or_404(project_id)
    if request.method == "POST":
        vote = Vote(
            user_id=current_user.id,
            project_id=project.id,
            ip_address=request.remote_addr,
        )
        db.session.add(vote)
        db.session.commit()
        flash('Your vote has been recorded!', 'success')
        return redirect(url_for("index"))
    metaData=g.metaData
    return render_template("vote.html", project=project, metaData=metaData)


@app.route("/get_project_data/<int:project_id>")
def get_project_data(project_id):
    project = Project.query.get_or_404(project_id)
    # Check if current user is the author of the project or the admin
    if (
        project.author == current_user.id or current_user.id == 1
    ):  # Assuming admin ID is 1
        return jsonify(project.to_dict())
    else:
        return jsonify({"error": "Unauthorized"}), 403


@app.route("/get_project_image/<filename>")
def get_project_image(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/update_project_data/<int:project_id>", methods=["POST"])
@login_required
def update_project_data(project_id):
    project = Project.query.get_or_404(project_id)
    if str(project.author) != str(current_user.id):
        return jsonify({"error": "Unauthorized"}), 403

    try:
        project.name = request.form.get("name", project.name)
        project.category = request.form.get("category", project.category)
        project.descriptionwhy = request.form["descriptionwhy"]
        project.public_benefit = request.form["public_benefit"]
        project.geoloc = request.form.get("geoloc", project.geoloc)

        # Handle image update if provided
        image_file = request.files.get("image_file")
        if image_file:
            filename = secure_filename(image_file.filename)
            image_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            image_file.save(image_path)
            project.image_file = filename

        db.session.commit()
        return jsonify({"success": "Project updated successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
